import datetime
import sqlite3
import os
from datetime import datetime
from mimesis import Person
import time
import openpyxl



option = Person('uk')
db = sqlite3.connect(os.path.join(os.getcwd(), 'base.db'))
cur = db.cursor()

# cur.execute("""CREATE TABLE IF NOT EXISTS test(
#         name text,
#         surname text,
#         fatherhood text,
#         date_birsday text,
#         date_death text,
#         sex text,
#         age text
#     )""")

# for _ in range(5000):
#     name = option.name()
#     last_name = option.last_name()
#     age = option.age()
#     telephone = option.telephone()
#     cur.execute('INSERT INTO traning VALUES (?, ?, ?, ?)', (name, last_name, age, telephone))
#     db.commit()

# def show_user():     # Перевіряє людей за ім'ям та прізвищем
#     name = input('Введіть ім\'я: ').strip().lower()
#     c = cur.execute(f'SELECT name, last_name, age, telephone FROM traning').fetchall()
#     for i in c:
#         if name in i[0].lower() or name in i[1]:
#             print(f'name: {i[0]}, last_name: {i[1]}, age: {i[2]}, telephone: {i[3]}')
# show_user()


# def export_to_sqlite():
#     '''Экспорт данных из xlsx в sqlite'''
#     # 2. Работа c xlsx файлом
#     # Читаем файл и лист1 книги excel
#     file_to_read = openpyxl.load_workbook('test.xlsx', data_only=True)
#     sheet = file_to_read['Аркуш1']
#
#     # Цикл по строкам начиная со второй (в первой заголовки)
#
#     for row in range(2, sheet.max_row):
#         # Объявление списка
#         data = []
#         # Цикл по столбцам от 1 до 4 ( 5 не включая)
#         for col in range(1, 8):
#             # value содержит значение ячейки с координатами row col
#             value = sheet.cell(row, col).value
#             # Список который мы потом будем добавлять
#             data.append(value)
#
#     # 3. Запись в базу и закрытие соединения
#
#         # Вставка данных в поля таблицы
#         cur.execute("INSERT INTO test VALUES (?, ?, ?, ?, ?, ?, ?);", (data[0], data[1], data[2], data[3], data[4], data[5], data[6]))
#
#     # сохраняем изменения
#     db.commit()
#
#
#
# def clear_base():
#     '''Очистка базы sqlite'''
#     cur.execute("DELETE FROM test")
#     db.commit()



# Запуск функции
# export_to_sqlite()
# clear_base()

# wook = ['.', ',', '/', '-', ' ']
# date = '20,13 2000'
# for i in wook:
#     if i in date:
#         date = date.replace(i, '.')
#
# print(date)

# def date_norm(self, date: str):
#     try:
#         self.date = date.strip()
#         if '/' in date:
#             self.date = list(map(int, self.date.split('/')))
#         elif '.' in self.date:
#             self.date = list(map(int, self.date.strip().split('.')))
#         elif '-' in self.date:
#             self.date = list(map(int, self.date.strip().split('-')))
#         elif ',' in self.date:
#             self.date = list(map(int, self.date.strip().split(',')))
#         else:
#             self.date = list(map(int, self.date.split()))
#         return self.date
#     except (IndexError, ValueError, TypeError):
#         print('Формат дати введений не правильно. Введіть дату ще раз.')
#         self.add_member()
