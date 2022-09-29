import sqlite3
import os
import datetime
from datetime import datetime
import openpyxl


db = sqlite3.connect(os.path.join(os.getcwd(), 'base.db'))
cur = db.cursor()


class People_base(object):

    def __init__(self):
        self.age = None
        self.sex = None
        self.date_death = None
        self.date_birsday = None
        self.fatherhood = None
        self.surname = None
        self.name = None

    def dell_user(self):
        self.name = input('Введіть ім\'я: ').strip().title()
        self.surname = input('Прізвище: ').strip().title()
        cur.execute(f'DELETE FROM people WHERE name = ? AND surname = ?', (self.name, self.surname))
        db.commit()
        print('Запис видалено.')
        print('--' * 50)

    def add_member(self):
        while True:
            self.name = input('Ім\'я: ').strip().title()
            if len(self.name) != 0:
                self.surname = input('Прізвище: ').strip().title()
                self.fatherhood = input('Ім\'я по батькові: ').strip().title()
                self.date_birsday = self.date_norm(input('Дата народження ДД ММ РРРР: '))
                self.date_death = self.date_norm(input('Дата смертів форматі ДД ММ РРРР: '))
                self.sex = input('Пол: ').strip()
                self.age = self.check_how_old(self.date_death)
                str_bir = '.'.join(map(str, self.date_birsday))
                str_ded = '.'.join(map(str, self.date_death))
                cur.execute('INSERT INTO people VALUES (?, ?, ?, ?, ?, ?, ?)',
                            (self.name, self.surname, self.fatherhood, str_bir, str_ded, self.sex, self.age))
                db.commit()
                print('Дані успішно внесені до бази даних')
                print('--' * 50)
                break
            else:
                print('Треба ввести ім\'я')
                print('--' * 50)

    def show_user(self):
        self.name = input('Введіть ім\'я: ').strip().lower()
        c = cur.execute(f'SELECT name, surname, fatherhood, date_birsday, date_death, sex FROM people').fetchall()
        for i in c:
            if self.name in i[0].lower() or self.name in i[1] or self.name in i[2]:
                print(
                    f'name: {i[0]}, surname: {i[1]}, fatherhood: {i[2]}, date_birsday: {i[3]}, date_death: {i[4]}, sex: {i[5]}')
                break
        else:
            print('Кандидатів немає')
            print('--' * 50)


    def date_norm(self, date: str):
        try:
            wook = ['.', ',', '/', '-', ' ']
            for i in wook:
                if i in date.strip():
                    date = date.replace(i, '.')
            self.date = list(map(int, date.split('.')))
            return self.date
        except (IndexError, ValueError, TypeError):
            print('Формат дати введений не правильно. Введіть дату ще раз.')
            self.add_member()

    def check_how_old(self, dead):
        borne = self.date_birsday
        if len(dead) == 0:
            date = str(datetime.today())[:10].split('-')
            result = list(map(int, date))
            today = result
            year = today[0] - borne[2]
            if borne[1] > today[1]:
                year -= 1
            elif borne[1] == today[1]:
                if borne[0] > today[2]:
                    year -= 1
            return year
        else:
            dead_day = self.date_death
            year = dead_day[2] - borne[2]
            if borne[1] > dead_day[1]:
                year -= 1
            elif borne[1] == dead_day[1]:
                if borne[0] > dead_day[0]:
                    year -= 1
            return year

    def export_to_db(self):
        try:
            while True:
                file_to_read = openpyxl.load_workbook(input('Вкажіть імя файлу: '), data_only=True)
                sheet = file_to_read[input('Вкажіть назву аркуша: ')]
                for row in range(2, sheet.max_row):
                    data = []
                    for col in range(1, 8):
                        value = sheet.cell(row, col).value
                        data.append(value)
                    cur.execute(f"INSERT INTO people VALUES (?, ?, ?, ?, ?, ?, ?);",
                                (data[0], data[1], data[2], data[3], data[4], data[5], data[6]))
                    db.commit()
                print('Дані вненесно успішно.')
                break
        except Exception:
            print('Не вірні дані спробуйте ще раз.')
            print('--' * 50)
            self.export_to_db()

    def clear_base(self):
        cur.execute("DELETE FROM people")
        db.commit()

